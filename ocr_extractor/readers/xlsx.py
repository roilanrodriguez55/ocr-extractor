"""XLSX reader — extract text directly from Excel 2007+ files.

Each worksheet becomes one ``=== PAGE <sheet name> ===`` /
``=== END PAGE <sheet name> ===`` marker block. Rows are emitted as
``" | col1 | col2 | col3"`` so the tabular structure survives in plain
text form.
"""

import re

from openpyxl import load_workbook
from ocr_extractor.result import wrap_page


def _clean_cell(value):
    """Stringify a cell value and collapse whitespace."""
    if value is None:
        return ""
    text = str(value)
    return re.sub(r"\s+", " ", text).strip()


def read_xlsx(path, *, dpi=300, lang="eng", verbose=True):
    """Extract text from a ``.xlsx`` file.

    Each worksheet produces one marker block, named after the sheet
    (so multi-sheet workbooks stay distinguishable). Each non-empty row
    becomes one line of the form ``" | col1 | col2 | col3"``.

    Parameters
    ----------
    path : str or pathlib.Path
        Path to the ``.xlsx`` file.
    dpi : int, optional
        Unused; kept for signature uniformity. Defaults to ``300``.
    lang : str, optional
        Unused; kept for signature uniformity. Defaults to ``"eng"``.
    verbose : bool, optional
        If ``True``, print a per-sheet progress line. Defaults to
        ``True``.

    Returns
    -------
    str
        Extracted text with one marker block per worksheet.
    """
    if verbose:
        print(f"Reading {path}...")

    # ``data_only=True`` returns computed values instead of formulas.
    # ``read_only=True`` is much faster and uses far less memory.
    wb = load_workbook(path, read_only=True, data_only=True)
    all_text = ""

    try:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if verbose:
                print(f"Processing sheet '{sheet_name}'...")

            lines = []
            for row in sheet.iter_rows(values_only=True):
                cells = [_clean_cell(v) for v in row]
                line = " | " + " | ".join(cells)
                stripped = line.strip(" |")
                if stripped:
                    lines.append(stripped)

            body = "\n".join(lines)
            all_text += wrap_page(sheet_name, body)
    finally:
        wb.close()

    return all_text


__all__ = ["read_xlsx"]