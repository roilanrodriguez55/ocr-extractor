"""Tests for the XLSX reader."""

import pytest
from openpyxl import Workbook

from ocr_extractor.readers.xlsx import read_xlsx


# ---------- synthetic XLSX fixtures ----------------------------------------


def _make_xlsx(path, sheets=None):
    """Build an XLSX where each sheet has rows of cell values.

    Parameters
    ----------
    path : pathlib.Path
        Where to save the .xlsx.
    sheets : dict[str, list[list]] or None
        Mapping of sheet name → 2D row list. If empty/None, a single
        empty sheet named ``"Sheet"`` is written (openpyxl requires at
        least one visible sheet to save).
    """
    wb = Workbook()
    sheet_names = list((sheets or {}).keys())
    if not sheet_names:
        # Keep the default sheet; just save.
        wb.save(str(path))
        return path

    first_name = sheet_names[0]
    first_ws = wb.active
    first_ws.title = first_name
    for row in sheets[first_name]:
        first_ws.append(row)
    for name in sheet_names[1:]:
        ws = wb.create_sheet(title=name)
        for row in sheets[name]:
            ws.append(row)

    wb.save(str(path))
    return path


# ---------- tests ----------------------------------------------------------


class TestReadXlsx:
    def test_empty_workbook_returns_empty(self, tmp_workdir):
        # openpyxl requires at least one sheet; we use the default
        # sheet (no rows) and expect only the marker block with no body.
        path = _make_xlsx(tmp_workdir / "empty.xlsx")
        text = read_xlsx(path, verbose=False)
        # One sheet → one marker block, but no row content.
        assert "=== PAGE Sheet ===" in text
        assert "=== END PAGE Sheet ===" in text
        # No actual content lines (just markers + whitespace).
        non_marker_lines = [
            l for l in text.split("\n")
            if l.strip() and not l.startswith("===")
        ]
        assert non_marker_lines == []

    def test_single_sheet_one_marker_block(self, tmp_workdir):
        path = _make_xlsx(
            tmp_workdir / "single.xlsx",
            sheets={"Sheet1": [["A", "B", "C"], [1, 2, 3]]},
        )
        text = read_xlsx(path, verbose=False)
        assert "=== PAGE Sheet1 ===" in text
        assert "=== END PAGE Sheet1 ===" in text
        assert "A | B | C" in text
        assert "1 | 2 | 3" in text
        assert "=== PAGE Sheet2 ===" not in text

    def test_multiple_sheets_each_marked(self, tmp_workdir):
        path = _make_xlsx(
            tmp_workdir / "multi.xlsx",
            sheets={
                "First": [["a"]],
                "Second": [["b"]],
                "Third": [["c"]],
            },
        )
        text = read_xlsx(path, verbose=False)
        for name in ("First", "Second", "Third"):
            assert f"=== PAGE {name} ===" in text
            assert f"=== END PAGE {name} ===" in text
        # Sheet order is preserved.
        assert text.index("PAGE First") < text.index("PAGE Second") < text.index("PAGE Third")

    def test_skips_completely_empty_rows(self, tmp_workdir):
        path = _make_xlsx(
            tmp_workdir / "with_empty.xlsx",
            sheets={"S": [["A", "B"], [None, None], ["", ""], ["X", "Y"]]},
        )
        text = read_xlsx(path, verbose=False)
        lines = [
            l for l in text.split("\n")
            if l.strip() and not l.startswith("===")
        ]
        # Two non-empty rows survive.
        assert lines == ["A | B", "X | Y"]

    def test_handles_none_cells(self, tmp_workdir):
        path = _make_xlsx(
            tmp_workdir / "with_none.xlsx",
            sheets={
                "S": [
                    ["A", None, "C"],
                    ["X", "B", None],
                ]
            },
        )
        text = read_xlsx(path, verbose=False)
        # None cells become empty strings; cells are still separated.
        assert "A |  | C" in text
        # "X | B" (trailing None cells become a trailing empty cell that
        # is then stripped, leaving "X | B").
        assert "X | B" in text

    def test_handles_numeric_cells(self, tmp_workdir):
        path = _make_xlsx(
            tmp_workdir / "numeric.xlsx",
            sheets={"S": [[1, 2.5, 3], [100, 200, 300]]},
        )
        text = read_xlsx(path, verbose=False)
        assert "1 | 2.5 | 3" in text
        assert "100 | 200 | 300" in text

    def test_preserves_strings_with_punctuation(self, tmp_workdir):
        path = _make_xlsx(
            tmp_workdir / "punct.xlsx",
            sheets={"S": [["Hello, world!", "Question?"]]},
        )
        text = read_xlsx(path, verbose=False)
        assert "Hello, world!" in text
        assert "Question?" in text

    def test_data_only_returns_computed_values(self, tmp_workdir):
        """openpyxl's data_only=True returns the cached value of a
        formula, not the formula string itself. Verify the reader passes
        ``data_only=True`` and ``read_only=True`` to ``load_workbook``.
        """
        path = _make_xlsx(
            tmp_workdir / "formulas.xlsx",
            sheets={"S": [[1, 2], [3, 4]]},
        )
        from unittest import mock
        from ocr_extractor.readers import xlsx as xlsx_mod

        with mock.patch.object(xlsx_mod, "load_workbook") as fake_load:
            # Build a real workbook as the return value, but only via the
            # unmocked loader so the call we care about is the single
            # ``fake_load`` invocation.
            real_wb = xlsx_mod.load_workbook.__wrapped__(path, read_only=True, data_only=True) \
                if hasattr(xlsx_mod.load_workbook, "__wrapped__") else None
            # Fallback: just inspect the kwargs passed to the mock.
            fake_load.return_value = mock.MagicMock()
            read_xlsx(path, verbose=False)
            fake_load.assert_called_once()
            _, kwargs = fake_load.call_args
            assert kwargs.get("data_only") is True
            assert kwargs.get("read_only") is True

    def test_verbose_prints_per_sheet(self, tmp_workdir, capsys):
        path = _make_xlsx(
            tmp_workdir / "v.xlsx",
            sheets={"First": [["a"]], "Second": [["b"]]},
        )
        read_xlsx(path, verbose=True)
        captured = capsys.readouterr()
        assert "First" in captured.out
        assert "Second" in captured.out

    def test_missing_file_raises(self, tmp_workdir):
        with pytest.raises(FileNotFoundError):
            read_xlsx(tmp_workdir / "nope.xlsx", verbose=False)